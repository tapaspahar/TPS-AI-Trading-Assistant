import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from core.database_manager import Database
from services.excel_report_exporter import ExcelReportExporter, REPORTS


class ExcelReportExporterTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.folder = Path(self.temp.name)
        self.db = Database(self.folder / "excel_reports.db")
        self.db.cursor.executemany(
            "INSERT INTO notifications (created_at, category, title, message, is_read, event_key) VALUES (?, ?, ?, ?, ?, ?)",
            [
                ("2026-08-12T10:00:00", "TEST", "Older", "Outside filter", 0, "old"),
                ("2026-08-13T10:00:00", "TEST", "Today", "Inside filter", 0, "new"),
            ],
        )
        self.db.cursor.execute(
            "INSERT INTO pcr_observations (captured_at, symbol, expiry, call_oi, put_oi, pcr_oi, call_volume, put_volume, pcr_volume, sentiment, confidence) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            ("2026-08-13T11:00:00", "NIFTY", "20-AUG-2026", 100, 120, 1.2, 50, 75, 1.5, "BULLISH", 80),
        )
        self.db.connection.commit()

    def tearDown(self):
        self.db.close()
        self.temp.cleanup()

    def test_single_report_honours_one_day_filter(self):
        destination = self.folder / "notifications.xlsx"
        counts = ExcelReportExporter(self.db).export(
            destination, ["notifications"], date(2026, 8, 13), date(2026, 8, 13)
        )
        self.assertEqual(counts, {"notifications": 1})
        workbook = load_workbook(destination)
        self.assertEqual(workbook.sheetnames, ["Notifications"])
        sheet = workbook["Notifications"]
        self.assertEqual(sheet.freeze_panes, "A5")
        self.assertTrue(sheet.auto_filter.ref)
        self.assertEqual(sheet.cell(5, 4).value, "Today")

    def test_all_reports_create_summary_and_every_sheet(self):
        destination = self.folder / "all_reports.xlsx"
        counts = ExcelReportExporter(self.db).export(destination)
        workbook = load_workbook(destination)
        self.assertEqual(workbook.sheetnames[0], "Export Summary")
        self.assertEqual(set(counts), {item.key for item in REPORTS})
        self.assertTrue({item.sheet for item in REPORTS}.issubset(set(workbook.sheetnames)))
        self.assertEqual(counts["notifications"], 2)
        self.assertEqual(counts["pcr_observations"], 1)

    def test_invalid_period_is_rejected(self):
        with self.assertRaisesRegex(ValueError, "Start date"):
            ExcelReportExporter(self.db).export(
                self.folder / "bad.xlsx", ["notifications"], date(2026, 8, 14), date(2026, 8, 13)
            )


if __name__ == "__main__":
    unittest.main()
