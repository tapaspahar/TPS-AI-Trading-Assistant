"""Excel export service for every persisted TPS report dataset."""

from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class ReportDefinition:
    key: str
    label: str
    sheet: str
    table: str
    date_column: str
    order_column: str


REPORTS = (
    ReportDefinition("trade_journal", "Trade Journal", "Trade Journal", "trades", "trade_date", "id"),
    ReportDefinition("auto_attempts", "Auto Trade Attempts", "Auto Attempts", "auto_trade_attempts", "trade_date", "id"),
    ReportDefinition("notifications", "Notification History", "Notifications", "notifications", "created_at", "id"),
    ReportDefinition("opportunities", "Auto Opportunities", "Opportunities", "auto_opportunities", "scanned_at", "id"),
    ReportDefinition("market_snapshots", "Market Snapshots", "Market Snapshots", "market_snapshots", "trade_date", "id"),
    ReportDefinition("pcr_observations", "PCR / OI Observations", "PCR and OI", "pcr_observations", "captured_at", "id"),
    ReportDefinition("gap_probability", "Gap Probability", "Gap Probability", "gap_probability_forecasts", "forecast_date", "id"),
    ReportDefinition("trend_memory", "Trend Memory", "Trend Memory", "daily_trend_memory", "trade_date", "id"),
    ReportDefinition("post_market", "Post Market Analysis", "Post Market Analysis", "post_market_tps_analysis", "trade_date", "id"),
    ReportDefinition("self_development", "AI Self-Development", "AI Self Development", "self_development_reviews", "trade_date", "id"),
)


class ExcelReportExporter:
    """Create styled, filterable XLSX reports from the TPS SQLite database."""

    def __init__(self, database):
        self.database = database

    @staticmethod
    def _as_date(value) -> date | None:
        if value in (None, ""):
            return None
        text = str(value).strip()
        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
        except ValueError:
            return None

    def _rows(self, definition: ReportDefinition, start: date | None, end: date | None) -> list[dict]:
        rows = self.database.cursor.execute(
            f'SELECT * FROM "{definition.table}" ORDER BY "{definition.order_column}" DESC'
        ).fetchall()
        selected = []
        for row in rows:
            item = dict(row)
            row_date = self._as_date(item.get(definition.date_column))
            if start and (row_date is None or row_date < start):
                continue
            if end and (row_date is None or row_date > end):
                continue
            selected.append(item)
        return selected

    @staticmethod
    def _display(value):
        if value is None:
            return ""
        if isinstance(value, str) and value[:1] in ("{", "["):
            try:
                parsed = json.loads(value)
                return json.dumps(parsed, ensure_ascii=False, indent=2)
            except (ValueError, TypeError):
                pass
        return value

    @staticmethod
    def _period_text(start: date | None, end: date | None) -> str:
        if not start and not end:
            return "All available records"
        if start == end:
            return start.strftime("%d-%m-%Y") if start else "All available records"
        return f"{start.strftime('%d-%m-%Y') if start else 'Beginning'} to {end.strftime('%d-%m-%Y') if end else 'Latest'}"

    def _write_report_sheet(self, workbook: Workbook, definition: ReportDefinition, rows: list[dict], period: str) -> None:
        ws = workbook.create_sheet(definition.sheet)
        columns = list(rows[0].keys()) if rows else [
            row["name"] for row in self.database.cursor.execute(f'PRAGMA table_info("{definition.table}")')
        ]
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(columns)))
        title = ws.cell(1, 1, f"TPS AI Trading Assistant — {definition.label}")
        title.font = Font(color="FFFFFF", bold=True, size=15)
        title.fill = PatternFill("solid", fgColor="14213D")
        title.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, len(columns)))
        ws.cell(2, 1, f"Period: {period} | Generated: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}")
        ws.cell(2, 1).font = Font(color="425466", italic=True)
        for index, column in enumerate(columns, 1):
            cell = ws.cell(4, index, column.replace("_", " ").title())
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = PatternFill("solid", fgColor="2563EB")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if rows:
            for row_index, row in enumerate(rows, 5):
                for column_index, column in enumerate(columns, 1):
                    cell = ws.cell(row_index, column_index, self._display(row.get(column)))
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    if row_index % 2 == 0:
                        cell.fill = PatternFill("solid", fgColor="EAF2FF")
        else:
            ws.cell(5, 1, "No records found for the selected period.")
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:{get_column_letter(max(1, len(columns)))}{max(5, 4 + len(rows))}"
        ws.row_dimensions[1].height = 26
        for index, column in enumerate(columns, 1):
            observed = [len(str(column))]
            observed.extend(len(str(self._display(row.get(column, "")))) for row in rows[:200])
            ws.column_dimensions[get_column_letter(index)].width = min(42, max(11, max(observed, default=11) + 2))

    def export(self, destination: str | Path, report_keys: Iterable[str] | None = None,
               start: date | None = None, end: date | None = None) -> dict[str, int]:
        if start and end and start > end:
            raise ValueError("Start date cannot be after end date.")
        selected_keys = set(report_keys or [item.key for item in REPORTS])
        definitions = [item for item in REPORTS if item.key in selected_keys]
        if not definitions:
            raise ValueError("Select at least one report to export.")
        workbook = Workbook()
        workbook.remove(workbook.active)
        period = self._period_text(start, end)
        counts = {}
        report_rows = []
        for definition in definitions:
            rows = self._rows(definition, start, end)
            counts[definition.key] = len(rows)
            report_rows.append((definition, rows))
        if len(definitions) > 1:
            summary = workbook.create_sheet("Export Summary")
            summary.append(["TPS AI Trading Assistant — Excel Report Center"])
            summary.append(["Selected period", period])
            summary.append(["Generated at", datetime.now().strftime("%d-%m-%Y %H:%M:%S")])
            summary.append([])
            summary.append(["Report", "Records"])
            for definition, rows in report_rows:
                summary.append([definition.label, len(rows)])
            summary["A1"].font = Font(color="FFFFFF", bold=True, size=15)
            summary["A1"].fill = PatternFill("solid", fgColor="14213D")
            for cell in summary[5]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = PatternFill("solid", fgColor="2563EB")
            summary.column_dimensions["A"].width = 34
            summary.column_dimensions["B"].width = 24
            summary.freeze_panes = "A6"
        for definition, rows in report_rows:
            self._write_report_sheet(workbook, definition, rows, period)
        output = Path(destination)
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output)
        return counts
